import os
import openpyxl

def parse_report(filepath):
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        
        if 'Executive Summary' in wb.sheetnames:
            ws_summary = wb['Executive Summary']
            summary_dict = {}
            for row in list(ws_summary.values)[1:]:
                if row and len(row) >= 2 and row[0]:
                    summary_dict[str(row[0])] = row[1]
            
            mapped_summary = {
                'Test Suite': 'Backend Security Verification',
                'Total Tests': summary_dict.get('Total verification checks run', 'N/A'),
                'Passed': summary_dict.get('✅ PASSED (fix confirmed in code)', 'N/A'),
                'Failed': summary_dict.get('❌ FAILED (fix missing/incorrect)', 'N/A'),
                'Pass Rate %': str(summary_dict.get('Pass Rate', 'N/A')).replace('%', ''),
                'Duration (sec)': 'N/A',
                'End Time': 'N/A'
            }
            
            ws_details = wb['Finding Results'] if 'Finding Results' in wb.sheetnames else wb['Detailed Check Results']
            detail_rows = list(ws_details.values)
            detail_headers = [str(h) for h in detail_rows[0]]
            details = []
            for r in detail_rows[1:]:
                if r and r[0] is not None:
                    d = dict(zip(detail_headers, r))
                    details.append({
                        'No.': d.get('Finding ID', '-'),
                        'Category': d.get('Category', '-'),
                        'Test Name': f"Severity: {d.get('Severity', '-')}",
                        'Status': 'PASSED' if 'PASS' in str(d.get('Overall Result', '')).upper() else 'FAILED'
                    })
            return mapped_summary, details

        # Parse Summary
        ws_summary = wb['Summary']
        rows = list(ws_summary.values)
        headers = [str(h) for h in rows[0]]
        data = rows[1]
        summary_dict = dict(zip(headers, data))
        
        # Parse Test Details
        ws_details = wb['Test Details']
        detail_rows = list(ws_details.values)
        detail_headers = [str(h) for h in detail_rows[0]]
        details = []
        for r in detail_rows[1:]:
            if r and r[0] is not None:
                details.append(dict(zip(detail_headers, r)))
            
        return summary_dict, details
    except Exception as e:
        print(f"Error parsing {filepath}: {e}")
        return {}, []

def main():
    # Configure UTF-8 stdout if possible to prevent Windows encoding crashes when printing emojis
    import sys
    if hasattr(sys.stdout, 'reconfigure'):
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')

    # Get the project root directory (assuming script is in /scripts)
    project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    
    frontend_path = os.path.join(project_root, "REPORTS", "app", "app seleium testing.xlsx")
    backend_path = os.path.join(project_root, "REPORTS", "app", "app security test.xlsx")
    
    frontend_summary, frontend_details = parse_report(frontend_path)
    backend_summary, backend_details = parse_report(backend_path)
    
    markdown_output = []
    markdown_output.append("# 🧪 SaathiCare Automated Test Verification Dashboard\n")
    markdown_output.append("This dashboard displays the test results verified from the completed test execution reports.\n")
    
    # Frontend Test Suite Summary
    if frontend_summary:
        markdown_output.append("## 💻 Frontend Test Suite Summary")
        markdown_output.append("| Metric | Value |")
        markdown_output.append("|---|---|")
        markdown_output.append(f"| **Test Suite** | {frontend_summary.get('Test Suite', 'Frontend Tests')} |")
        markdown_output.append(f"| **Total Test Cases** | {frontend_summary.get('Total Tests', len(frontend_details))} |")
        markdown_output.append(f"| **Passed** | ✅ {frontend_summary.get('Passed', 'N/A')} |")
        markdown_output.append(f"| **Failed** | ❌ {frontend_summary.get('Failed', 'N/A')} |")
        markdown_output.append(f"| **Pass Rate** | **{frontend_summary.get('Pass Rate %', 'N/A')}%** |")
        markdown_output.append(f"| **Duration** | {frontend_summary.get('Duration (sec)', 'N/A')} sec |")
        markdown_output.append(f"| **Timestamp** | {frontend_summary.get('End Time', 'N/A')} |")
        markdown_output.append("\n")
    
    # Backend Summary
    if backend_summary:
        markdown_output.append("## ⚙️ Backend Test Verification Summary")
        markdown_output.append("| Metric | Value |")
        markdown_output.append("|---|---|")
        markdown_output.append(f"| **Test Suite** | {backend_summary.get('Test Suite', 'Backend Tests')} |")
        markdown_output.append(f"| **Total Test Cases** | {backend_summary.get('Total Tests', len(backend_details))} |")
        markdown_output.append(f"| **Passed** | ✅ {backend_summary.get('Passed', 'N/A')} |")
        markdown_output.append(f"| **Failed** | ❌ {backend_summary.get('Failed', 'N/A')} |")
        markdown_output.append(f"| **Pass Rate** | **{backend_summary.get('Pass Rate %', 'N/A')}%** |")
        markdown_output.append(f"| **Duration** | {backend_summary.get('Duration (sec)', 'N/A')} sec |")
        markdown_output.append(f"| **Timestamp** | {backend_summary.get('End Time', 'N/A')} |")
        markdown_output.append("\n")
    
    # Frontend Details Expandable Section
    if frontend_details:
        markdown_output.append(f"### 📋 Frontend Test Cases Detail Breakdowns")
        markdown_output.append(f"<details><summary>Click to view all Frontend Test Cases ({len(frontend_details)} tests)</summary>\n")
        markdown_output.append("| No. | Category | Test Name | Status |")
        markdown_output.append("|---|---|---|---|")
        for r in frontend_details:
            status_emoji = "✅ PASSED" if str(r.get("Status")).upper() == "PASSED" else "❌ FAILED"
            markdown_output.append(f"| {r.get('No.', '-')} | {r.get('Category', '-')} | `{r.get('Test Name', '-')}` | {status_emoji} |")
        markdown_output.append("\n</details>\n")
    
    # Backend Details Expandable Section
    if backend_details:
        markdown_output.append(f"### 🔐 Backend Test Cases Detail Breakdowns")
        markdown_output.append(f"<details><summary>Click to view all Backend Test Cases ({len(backend_details)} tests)</summary>\n")
        markdown_output.append("| No. | Category | Test Name | Status |")
        markdown_output.append("|---|---|---|---|")
        for r in backend_details:
            status_emoji = "✅ PASSED" if str(r.get("Status")).upper() == "PASSED" else "❌ FAILED"
            markdown_output.append(f"| {r.get('No.', '-')} | {r.get('Category', '-')} | `{r.get('Test Name', '-')}` | {status_emoji} |")
        markdown_output.append("\n</details>\n")
    
    markdown_output.append("## 📦 Downloadable Test Report Artifacts")
    markdown_output.append("The full Excel spreadsheets (`.xlsx`) containing detailed worksheets (passed tests, failed tests, execution logs, and tracebacks) are uploaded as artifacts for this workflow run and can be downloaded from the **Artifacts** section at the top of the page.")
    
    full_markdown = "\n".join(markdown_output)
    
    # Write to GITHUB_STEP_SUMMARY
    summary_file = os.environ.get("GITHUB_STEP_SUMMARY")
    if summary_file:
        with open(summary_file, "w", encoding="utf-8") as f:
            f.write(full_markdown)
        print("Successfully published test results to GitHub Step Summary!")
    else:
        print(full_markdown)

if __name__ == "__main__":
    main()
