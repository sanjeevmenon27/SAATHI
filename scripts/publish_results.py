import os
import openpyxl

def parse_report(filepath):
    try:
        if not os.path.exists(filepath):
            print(f"File not found: {filepath}")
            return {}, []
        wb = openpyxl.load_workbook(filepath, data_only=True)
        
        summary_dict = {}
        details = []

        if 'Summary' in wb.sheetnames and 'Test Details' in wb.sheetnames:
            ws_summary = wb['Summary']
            rows = list(ws_summary.values)
            if len(rows) >= 2:
                headers = [str(h) for h in rows[0]]
                data = rows[1]
                summary_dict = dict(zip(headers, data))
                
            ws_details = wb['Test Details']
            detail_rows = list(ws_details.values)
            if len(detail_rows) >= 1:
                detail_headers = [str(h) for h in detail_rows[0]]
                for r in detail_rows[1:]:
                    if r and r[0] is not None:
                        details.append(dict(zip(detail_headers, r)))
                        
        elif 'Severity Summary' in wb.sheetnames and 'Vulnerability Findings' in wb.sheetnames:
            ws_summary = wb['Severity Summary']
            rows = list(ws_summary.values)
            total = 0
            for r in rows[1:]:
                if r and r[0] == 'Total':
                    total = r[1]
            summary_dict = {
                'Test Suite': 'Vulnerability Scan',
                'Total Tests': total,
                'Passed': 'N/A',
                'Failed': 'N/A',
                'Pass Rate %': 'N/A',
                'Duration (sec)': 'N/A',
                'End Time': 'N/A'
            }
            
            ws_details = wb['Vulnerability Findings']
            detail_rows = list(ws_details.values)
            if len(detail_rows) >= 1:
                detail_headers = [str(h) for h in detail_rows[0]]
                for r in detail_rows[1:]:
                    if r and r[0] is not None:
                        d = dict(zip(detail_headers, r))
                        d['Test Name'] = d.get('Vulnerability Type', '')
                        details.append(d)

        elif 'Summary by Severity' in wb.sheetnames and 'Finding Results' in wb.sheetnames:
            ws_summary = wb['Summary by Severity']
            rows = list(ws_summary.values)
            total = 0
            passed = 0
            failed = 0
            for r in rows[1:]:
                if r and r[0] == 'TOTAL':
                    total = r[1]
                    passed = r[2]
                    failed = r[3]
            
            summary_dict = {
                'Test Suite': 'Security Scan',
                'Total Tests': total,
                'Passed': passed,
                'Failed': failed,
                'Pass Rate %': round((passed / total * 100) if total else 0, 2),
                'Duration (sec)': 'N/A',
                'End Time': 'N/A'
            }

            ws_details = wb['Finding Results']
            detail_rows = list(ws_details.values)
            if len(detail_rows) >= 1:
                detail_headers = [str(h) for h in detail_rows[0]]
                for r in detail_rows[1:]:
                    if r and r[0] is not None:
                        d = dict(zip(detail_headers, r))
                        d['Test Name'] = str(d.get('Category', '')) + ' Check'
                        d['Status'] = d.get('Overall Result', '')
                        details.append(d)
        else:
            print(f"Unknown format in {filepath}")
        
        return summary_dict, details
    except Exception as e:
        print(f"Error parsing {filepath}: {e}")
        return {}, []

def add_section(markdown_output, summary, details, title, icon):
    if not summary:
        return
    markdown_output.append(f"## {icon} {title} Summary")
    markdown_output.append("| Metric | Value |")
    markdown_output.append("|---|---|")
    markdown_output.append(f"| **Test Suite** | {summary.get('Test Suite', title)} |")
    markdown_output.append(f"| **Total Test Cases** | {summary.get('Total Tests', len(details))} |")
    markdown_output.append(f"| **Passed** | ✅ {summary.get('Passed', 'N/A')} |")
    markdown_output.append(f"| **Failed** | ❌ {summary.get('Failed', 'N/A')} |")
    markdown_output.append(f"| **Pass Rate** | **{summary.get('Pass Rate %', 'N/A')}%** |")
    markdown_output.append(f"| **Duration** | {summary.get('Duration (sec)', 'N/A')} sec |")
    markdown_output.append(f"| **Timestamp** | {summary.get('End Time', 'N/A')} |")
    markdown_output.append("\n")
    
    if details:
        markdown_output.append(f"### 📋 {title} Detail Breakdowns")
        markdown_output.append(f"<details><summary>Click to view all {title} Cases ({len(details)} tests)</summary>\n")
        markdown_output.append("| No. | Category | Test Name | Status |")
        markdown_output.append("|---|---|---|---|")
        for r in details:
            status_val = str(r.get("Status", r.get("Overall Result", ""))).upper()
            if "PASS" in status_val:
                status_emoji = "✅ PASSED"
            elif "FAIL" in status_val:
                status_emoji = "❌ FAILED"
            else:
                status_emoji = status_val if status_val.strip() else "⚠️ N/A"
            
            finding_id = r.get('No.', r.get('Finding ID', '-'))
            markdown_output.append(f"| {finding_id} | {r.get('Category', '-')} | `{r.get('Test Name', '-')}` | {status_emoji} |")
        markdown_output.append("\n</details>\n")

def main():
    import sys
    if hasattr(sys.stdout, 'reconfigure'):
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')

    project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    
    reports = [
        ("Mobile App Security Test", "app security test.xlsx", "🔐"),
        ("Mobile App Selenium Testing", "app seleium testing.xlsx", "📱"),
        ("Backend Saathi Care Result", "backend saati care result.xlsx", "⚙️"),
        ("Frontend Saathi Care Result", "frontend saathi care result.xlsx", "💻")
    ]
    
    markdown_output = []
    markdown_output.append("# 🧪 SaathiCare Full Test Verification Dashboard\n")
    markdown_output.append("This dashboard displays the test results verified from the completed test execution reports for the entire system.\n")
    
    for title, filename, icon in reports:
        filepath = os.path.join(project_root, filename)
        summary, details = parse_report(filepath)
        add_section(markdown_output, summary, details, title, icon)

    markdown_output.append("## 📦 Downloadable Test Report Artifacts")
    markdown_output.append("The full Excel spreadsheets (`.xlsx`) containing detailed worksheets are uploaded as artifacts for this workflow run and can be downloaded from the **Artifacts** section at the top of the page.")
    
    full_markdown = "\n".join(markdown_output)
    
    summary_file = os.environ.get("GITHUB_STEP_SUMMARY")
    if summary_file:
        with open(summary_file, "w", encoding="utf-8") as f:
            f.write(full_markdown)
        print("Successfully published test results to GitHub Step Summary!")
    else:
        print(full_markdown)

if __name__ == "__main__":
    main()
