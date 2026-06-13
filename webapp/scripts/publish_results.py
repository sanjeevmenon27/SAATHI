import os
import openpyxl

def parse_report(filepath):
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        summary_dict = {}
        details = []

        if 'Summary' in wb.sheetnames and 'Test Details' in wb.sheetnames:
            ws_summary = wb['Summary']
            rows = list(ws_summary.values)
            headers = [str(h) for h in rows[0]]
            data = rows[1]
            summary_dict = dict(zip(headers, data))
            
            ws_details = wb['Test Details']
            detail_rows = list(ws_details.values)
            detail_headers = [str(h) for h in detail_rows[0]]
            for r in detail_rows[1:]:
                if r and r[0] is not None:
                    details.append(dict(zip(detail_headers, r)))
        
        elif 'Severity Summary' in wb.sheetnames and 'Vulnerability Findings' in wb.sheetnames:
            ws_summary = wb['Severity Summary']
            rows = list(ws_summary.values)
            # Backend severity summary looks like: [('Severity', 'Count'), ('Critical', 4), ...]
            summary_dict = {'Test Suite': 'Backend Security & Vulnerabilities Tests'}
            passed_count = 0
            failed_count = 0
            total_count = 0
            
            ws_details = wb['Vulnerability Findings']
            detail_rows = list(ws_details.values)
            detail_headers = [str(h) for h in detail_rows[0]]
            
            for r in detail_rows[1:]:
                if r and r[0] is not None:
                    row_dict = dict(zip(detail_headers, r))
                    details.append(row_dict)
                    
                    status = str(row_dict.get('Status', '')).upper()
                    if 'PASS' in status:
                        passed_count += 1
                    elif 'FAIL' in status:
                        failed_count += 1
                    total_count += 1
                    
            summary_dict['Total Tests'] = total_count
            summary_dict['Passed'] = passed_count
            summary_dict['Failed'] = failed_count
            if total_count > 0:
                summary_dict['Pass Rate %'] = round((passed_count / total_count) * 100, 2)
            else:
                summary_dict['Pass Rate %'] = 0
            summary_dict['Duration (sec)'] = 'N/A'
            summary_dict['End Time'] = 'N/A'

        return summary_dict, details
    except Exception as e:
        print(f"Error parsing {filepath}: {e}")
        return {}, []

def main():
    import sys
    if hasattr(sys.stdout, 'reconfigure'):
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')

    project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    
    frontend_path = os.path.join(project_root, "REPORTS", "frontend", "frontend saathi care result.xlsx")
    backend_path = os.path.join(project_root, "REPORTS", "backend", "backend saati care result.xlsx")
    
    frontend_summary, frontend_details = parse_report(frontend_path)
    backend_summary, backend_details = parse_report(backend_path)
    
    markdown_output = []
    markdown_output.append("# 🧪 SaathiCare Automated Test Verification Dashboard\n")
    markdown_output.append("This dashboard displays the test results verified from the completed test execution reports.\n")
    
    # Frontend Test Suite Summary
    if frontend_summary:
        markdown_output.append("## 💻 Frontend Test Suite Summary")
        markdown_output.append("| Metric | Value |")
        markdown_output.append("|---|---|")
        markdown_output.append(f"| **Test Suite** | {frontend_summary.get('Test Suite', 'Frontend Tests')} |")
        markdown_output.append(f"| **Total Test Cases** | {frontend_summary.get('Total Tests', len(frontend_details))} |")
        markdown_output.append(f"| **Passed** | ✅ {frontend_summary.get('Passed', 'N/A')} |")
        markdown_output.append(f"| **Failed** | ❌ {frontend_summary.get('Failed', 'N/A')} |")
        markdown_output.append(f"| **Pass Rate** | **{frontend_summary.get('Pass Rate %', 'N/A')}%** |")
        markdown_output.append(f"| **Duration** | {frontend_summary.get('Duration (sec)', 'N/A')} sec |")
        markdown_output.append(f"| **Timestamp** | {frontend_summary.get('End Time', 'N/A')} |")
        markdown_output.append("\n")
    
    # Backend Summary
    if backend_summary:
        markdown_output.append("## ⚙️ Backend Test Verification Summary")
        markdown_output.append("| Metric | Value |")
        markdown_output.append("|---|---|")
        markdown_output.append(f"| **Test Suite** | {backend_summary.get('Test Suite', 'Backend Tests')} |")
        markdown_output.append(f"| **Total Test Cases** | {backend_summary.get('Total Tests', len(backend_details))} |")
        markdown_output.append(f"| **Passed** | ✅ {backend_summary.get('Passed', 'N/A')} |")
        markdown_output.append(f"| **Failed** | ❌ {backend_summary.get('Failed', 'N/A')} |")
        markdown_output.append(f"| **Pass Rate** | **{backend_summary.get('Pass Rate %', 'N/A')}%** |")
        markdown_output.append(f"| **Duration** | {backend_summary.get('Duration (sec)', 'N/A')} sec |")
        markdown_output.append(f"| **Timestamp** | {backend_summary.get('End Time', 'N/A')} |")
        markdown_output.append("\n")
    
    # Frontend Details Expandable Section
    if frontend_details:
        markdown_output.append(f"### 📋 Frontend Test Cases Detail Breakdowns")
        markdown_output.append(f"<details><summary>Click to view all Frontend Test Cases ({len(frontend_details)} tests)</summary>\n")
        markdown_output.append("| No. | Category | Test Name | Status |")
        markdown_output.append("|---|---|---|---|")
        for r in frontend_details:
            status_emoji = "✅ PASSED" if "PASS" in str(r.get("Status")).upper() else "❌ FAILED"
            markdown_output.append(f"| {r.get('No.', '-')} | {r.get('Category', '-')} | `{r.get('Test Name', '-')}` | {status_emoji} |")
        markdown_output.append("\n</details>\n")
    
    # Backend Details Expandable Section
    if backend_details:
        markdown_output.append(f"### 🔐 Backend Test Cases Detail Breakdowns")
        markdown_output.append(f"<details><summary>Click to view all Backend Test Cases ({len(backend_details)} tests)</summary>\n")
        markdown_output.append("| ID | Category | Description | Status |")
        markdown_output.append("|---|---|---|---|")
        for r in backend_details:
            status_emoji = "✅ PASSED" if "PASS" in str(r.get("Status")).upper() else "❌ FAILED"
            id_val = r.get('ID', r.get('No.', '-'))
            cat_val = r.get('Category', '-')
            desc_val = r.get('Description', r.get('Test Name', '-'))
            markdown_output.append(f"| {id_val} | {cat_val} | `{desc_val}` | {status_emoji} |")
        markdown_output.append("\n</details>\n")
    
    markdown_output.append("## 📦 Downloadable Test Report Artifacts")
    markdown_output.append("The full Excel spreadsheets (`.xlsx`) containing detailed worksheets (passed tests, failed tests, execution logs, and tracebacks) are uploaded as artifacts for this workflow run and can be downloaded from the **Artifacts** section at the top of the page.")
    
    full_markdown = "\n".join(markdown_output)
    
    # Write to GITHUB_STEP_SUMMARY
    summary_file = os.environ.get("GITHUB_STEP_SUMMARY")
    if summary_file:
        with open(summary_file, "w", encoding="utf-8") as f:
            f.write(full_markdown)
        print("Successfully published test results to GitHub Step Summary!")
    else:
        print(full_markdown)

if __name__ == "__main__":
    main()
